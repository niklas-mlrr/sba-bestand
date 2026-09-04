"""Synthetisches Workbook und Fake-IServ - fuer Tests beider Repos.

Wird von ``sba-bestand`` und ``sba-dashboard`` benutzt: beide sollen gegen
*dieselbe* nachgebaute Struktur testen, sonst driften die Annahmen auseinander.
Das Blatt bildet die vier gemessenen Eigenheiten des echten Bestandsblatts nach:

  1. Fachblock = vier Spalten (Angemeldet | Bestand | Bestellt | zu bestellen)
  2. "zu bestellen" ist eine echte Formel, kein Wert
  3. Mehrjahresbaende: Bestand/Bestellt/zu-bestellen sind ueber mehrere
     Jahrgangszeilen verbunden, Angemeldet bleibt je Jahrgang einzeln
  4. Ein Merge ueber die volle Blockbreite = "nicht angeboten"-Sperrflaeche

Kein Netz, keine echte Excel-Datei, kein pytest-Import.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableFormula, TableStyleInfo

SHEET_NAME = "Bestand- und Nachbestellung"
ZUSTAENDE = ("Angemeldet", "Bestand", "Bestellt", "zu bestellen")

# Faecher des Sek-I-Blocks, in Spaltenreihenfolge ab Spalte B.
BLOCKS_SEK1 = ("Deutsch", "Erdkunde", "Latein")
# Zweiter Fach-Block (Oberstufe): leeres Label faellt auf die Zeile-1-Beschriftung
# zurueck, genau wie im echten Blatt (B11:E11 ist leer -> "Deutsch").
BLOCKS_SEK2 = ("", "Erdkunde (eA)", "")

ISBN_DEUTSCH_5 = "9783062052224"
ISBN_DEUTSCH_6 = "9783062052231"
ISBN_DEUTSCH_7 = "9783062052248"
ISBN_ERDKUNDE_56 = "9783121052073"
ISBN_ERDKUNDE_7 = "9783121052080"
ISBN_DEUTSCH_12 = "9783140282680"
ISBN_ERDKUNDE_EA_12 = "9783507533059"


def _block_start(index: int) -> int:
    """Erste Spalte (1-basiert) des index-ten Fachblocks; Block 0 beginnt bei B."""
    return 2 + index * 4


def _build_raster(ws) -> None:
    ws["A1"] = "Fach"
    ws["A2"] = "Zustand"
    for i, fach in enumerate(BLOCKS_SEK1):
        c = _block_start(i)
        ws.cell(1, c).value = fach
        ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c + 3)
        for j, zustand in enumerate(ZUSTAENDE):
            ws.cell(2, c + j).value = zustand

    # Jahrgangszeilen 3-5 = Jg 5, 6, 7
    for row, grade in ((3, 5), (4, 6), (5, 7)):
        ws.cell(row, 1).value = f"Jahrgang {grade}"

    # Deutsch: jeder Jahrgang ein eigenes Buch, keine Verbuende.
    for row in (3, 4, 5):
        ws.cell(row, 5).value = f"=B{row}-C{row}-D{row}"

    # Erdkunde: ein Band fuer Jg 5-6 -> Bestand/Bestellt/zu bestellen verbunden,
    # Angemeldet bleibt je Jahrgang einzeln und wird in der Formel summiert.
    for col in (7, 8, 9):
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
    ws["I3"] = "=F3+F4-G3-H3"
    ws["I5"] = "=F5-G5-H5"

    # Latein: in Jg 5 und 6 nicht angeboten -> Merge ueber die volle Blockbreite.
    ws.merge_cells(start_row=3, start_column=10, end_row=4, end_column=13)
    ws["M5"] = "=J5-K5-L5"

    # Zweiter Fach-Block (Oberstufe), ohne eigene Zustand-Zeile - wie im Original.
    ws["A7"] = "Fach"
    for i, fach in enumerate(BLOCKS_SEK2):
        c = _block_start(i)
        if fach:
            ws.cell(7, c).value = fach
        ws.merge_cells(start_row=7, start_column=c, end_row=7, end_column=c + 3)
    ws["A8"] = "Jahrgang 12"
    ws["E8"] = "=B8-C8-D8"
    ws["I8"] = "=F8-G8-H8"
    # Latein auch in Jg 12 nicht angeboten.
    ws.merge_cells(start_row=8, start_column=10, end_row=8, end_column=13)

    ws["A10"] = "Stand"
    ws.merge_cells(start_row=10, start_column=2, end_row=10, end_column=8)


def _build_bestellt(ws) -> None:
    ws.append(["B. Lfd.-Nr.", "Lfd. Nr.", "Stückzahl", "Verfasser/Titel", "Verlag",
               "Best.-Nr. (ISBN)", "Einzelpreis (brutto)"])
    ws.append([8, 1, 10, "Deutschbuch 6", "Cornelsen", "978-3-06-205223-1", 20.0])
    ws.append([8, 2, 5, "Deutschbuch 6", "Cornelsen", "978-3-06-205223-1", 20.0])
    ws.append([9, 1, 30, "Terra 5/6", "Klett", "978-3-12-105207-3", 25.0])


_ZU_BESTELLEN_COLS = (
    ("Bestell Nr.", None, "Ergebnis"),
    ("Jahrgang", None, None),
    ("Fach", None, None),
    ("Stückzahl", "sum", None),
    ("Verfasser/Titel", None, None),
    ("Verlag", None, None),
    ("Best.-Nr. (ISBN)", None, None),
    ("Einzelpreis (brutto)", None, None),
    ("Gesamtpreis (brutto)", "sum", None),
)


def _build_zu_bestellen(ws) -> None:
    for i, (name, _fn, _label) in enumerate(_ZU_BESTELLEN_COLS, start=1):
        ws.cell(1, i).value = name
    # Eine Alt-Datenzeile: ihre Bestell-Nr. muss beim Neuaufbau per ISBN
    # wieder an dieselbe Buchreihe wandern.
    ws["A2"] = 42
    ws["G2"] = "978-3-12-105207-3"
    ws["A3"] = "Ergebnis"

    columns = [
        TableColumn(id=i, name=name, totalsRowFunction=fn, totalsRowLabel=label)
        for i, (name, fn, label) in enumerate(_ZU_BESTELLEN_COLS, start=1)
    ]
    columns[-1].calculatedColumnFormula = TableFormula(
        attr_text=(
            'IF(OR(zuBestellen[[#This Row],[Stückzahl]]="",'
            'zuBestellen[[#This Row],[Einzelpreis (brutto)]]=""),"",'
            'zuBestellen[[#This Row],[Stückzahl]]*zuBestellen[[#This Row],[Einzelpreis (brutto)]])'
        )
    )
    table = Table(displayName="zuBestellen", name="zuBestellen", ref="A1:I3",
                  headerRowCount=1, totalsRowCount=1, tableColumns=columns)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    table.autoFilter = AutoFilter(ref="A1:I2")
    ws.add_table(table)


def build_workbook(path: Path) -> Path:
    """Schreibt das synthetische Workbook nach path und gibt den Pfad zurueck."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _build_raster(ws)
    _build_bestellt(wb.create_sheet("bestellt"))
    _build_zu_bestellen(wb.create_sheet("zu Bestellen"))
    wb.save(path)
    return path


# ── Fake-IServ ────────────────────────────────────────────────────────────────

class _Series:
    def __init__(self, isbn, title, total, publisher, price):
        self.isbn, self.title, self.total = isbn, title, total
        self.publisher, self.price = publisher, price


_BOOKS = {
    5:  [(ISBN_DEUTSCH_5, "Deutschbuch 5", ["Deutsch"]),
         (ISBN_ERDKUNDE_56, "Terra 5/6", ["Erdkunde"])],
    6:  [(ISBN_DEUTSCH_6, "Deutschbuch 6", ["Deutsch"]),
         (ISBN_ERDKUNDE_56, "Terra 5/6", ["Erdkunde"])],
    7:  [(ISBN_DEUTSCH_7, "Deutschbuch 7", ["Deutsch"]),
         (ISBN_ERDKUNDE_7, "Terra 7/8", ["Erdkunde"])],
    12: [(ISBN_DEUTSCH_12, "Deutschbuch Oberstufe", ["Deutsch"]),
         (ISBN_ERDKUNDE_EA_12, "Terra Erhöhtes Anforderungsniveau", ["Erdkunde"])],
}

_SERIES = {
    ISBN_DEUTSCH_5: _Series(ISBN_DEUTSCH_5, "Deutschbuch 5", 40, "Cornelsen", 22.5),
    ISBN_DEUTSCH_6: _Series(ISBN_DEUTSCH_6, "Deutschbuch 6", 30, "Cornelsen", 20.0),
    ISBN_DEUTSCH_7: _Series(ISBN_DEUTSCH_7, "Deutschbuch 7", 90, "Cornelsen", 21.0),
    ISBN_ERDKUNDE_56: _Series(ISBN_ERDKUNDE_56, "Terra 5/6", 60, "Klett", 25.0),
    ISBN_ERDKUNDE_7: _Series(ISBN_ERDKUNDE_7, "Terra 7/8", 55, "Klett", 26.0),
    ISBN_DEUTSCH_12: _Series(ISBN_DEUTSCH_12, "Deutschbuch Oberstufe", 20, "Cornelsen", 30.0),
    ISBN_ERDKUNDE_EA_12: _Series(ISBN_ERDKUNDE_EA_12, "Terra Erhöhtes Anforderungsniveau",
                                 10, "Klett", 32.0),
}

# (Jahrgang, ISBN) -> (Anmeldungen, davon bezahlt)
_ENROLLMENTS = {
    (5, ISBN_DEUTSCH_5): (50, 40),
    (5, ISBN_ERDKUNDE_56): (48, 30),
    (6, ISBN_DEUTSCH_6): (45, 45),
    (6, ISBN_ERDKUNDE_56): (44, 20),
    (7, ISBN_DEUTSCH_7): (52, 10),
    (7, ISBN_ERDKUNDE_7): (51, 51),
    (12, ISBN_DEUTSCH_12): (25, 25),
    (12, ISBN_ERDKUNDE_EA_12): (18, 18),
}


class _Schoolyears:
    def get_current(self):
        return {"id": "2026/2027"}

    def get_booklists(self, sy_id):
        return [{"id": 100 + g, "grade": g} for g in sorted(_BOOKS)]

    def get_booklist(self, sy_id, bl_id):
        grade = bl_id - 100
        items = [
            {"borrowable": True, "series": isbn,
             "series_data": {"isbn": isbn, "title": title, "fee": 5.0, "subjectsFlat": subjects}}
            for isbn, title, subjects in _BOOKS.get(grade, [])
        ]
        # Ein nicht ausleihbarer eBook-Eintrag, der herausgefiltert werden muss.
        items.append({"borrowable": True, "series": "9783128640105",
                      "series_data": {"isbn": "9783128640105", "title": "Deutschbuch eBook",
                                      "fee": 5.0, "subjectsFlat": ["Deutsch"]}})
        return {"sections": [{"options": [{"items": items}]}]}


class _Admin:
    def get_enrollments(self, sy_id):
        out = []
        for (grade, isbn), (count, paid) in _ENROLLMENTS.items():
            for i in range(count):
                out.append({
                    "deleted_at": None,
                    "Booklist": {"grade": grade},
                    "amountOpen": 0 if i < paid else 12.0,
                    "booklistItems": [{"series": isbn}],
                })
        # Eine geloeschte Anmeldung darf nicht mitzaehlen.
        out.append({"deleted_at": "2026-01-01", "Booklist": {"grade": 5},
                    "amountOpen": 0, "booklistItems": [{"series": ISBN_DEUTSCH_5}]})
        return out


class _SeriesApi:
    def get_all(self, detailed=False):
        return list(_SERIES.values())


class FakeClient:
    """Minimaler Ersatz fuer AusleiheClient - nur die genutzten Endpunkte."""

    def __init__(self, *args, **kwargs):
        self.schoolyears = _Schoolyears()
        self.admin = _Admin()
        self.series = _SeriesApi()

    def login(self):
        return True
