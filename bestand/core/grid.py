"""Struktur des Bestandsblatts - liest das Raster, ohne je ins Netz zu gehen.

Vier gemessene Eigenheiten des echten Workbooks bestimmen dieses Modul:

1. Ein Fachblock ist vier Spalten breit: Angemeldet | Bestand | Bestellt |
   zu bestellen. Eine ``Bezahlt``-Spalte gibt es nicht (mehr); sie bleibt in
   :data:`WRITABLE_ZUSTAENDE`, damit aeltere Mappen weiterhin laufen.
2. ``zu bestellen`` sind echte Formeln. Das Workbook wird deshalb immer mit
   ``data_only=False`` geladen; wer den Wert braucht, rechnet ihn selbst.
3. Die Merge-Topologie der **Bestand**-Spalte definiert die Zeilengruppe
   (K3:K4 = Jg 5-6, BC4:BC7 = Jg 6-9, AU9:AU12 = Jg 11-12). Die
   Angemeldet-Spalte bleibt je Jahrgang einzeln und wird summiert.
4. Ein Merge ueber die **volle Blockbreite** in Jahrgangszeilen ist eine
   "nicht angeboten"-Sperrflaeche (R3:U5, AD3:AG6, B12:E13 ...). Solche
   Flaechen werden nie beschrieben und tauchen in keiner Liste auf.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Callable, Optional

from openpyxl.utils import get_column_letter

# Zustaende, in die geschrieben werden darf, in Spaltenreihenfolge des Blocks.
WRITABLE_ZUSTAENDE = ("angemeldet", "bezahlt", "bestand", "bestellt")
# Der Zustand, der einen neuen Fachblock eroeffnet.
BLOCK_START_ZUSTAND = "angemeldet"

SKIP_NO_ZUSTAND = "kein Zustand-Label"
SKIP_ZUSTAND_NOT_WRITABLE = "Zustand nicht schreibbar"
SKIP_NO_FACH = "kein Fach-Label"
SKIP_NOT_OFFERED = "nicht angeboten"

Debug = Optional[Callable[[str], None]]


# ── Zellen-Hilfsfunktionen ────────────────────────────────────────────────────

def merge_deckt_ab(merged, row: int, col: int) -> bool:
    """Ob ``merged`` die Zelle (row, col) ueberdeckt - reiner Ganzzahlvergleich.

    Der naheliegende Weg waere ``f"{get_column_letter(col)}{row}" in merged``.
    Er ist korrekt, aber unbrauchbar langsam: ``CellRange.__contains__`` baut
    aus dem Text jedes Mal ein neues ``CellRange``, und dessen Konstruktor
    laesst vier openpyxl-Deskriptoren ihre Werte pruefen. Gemessen am echten
    Blatt (2026-09-04, 142 Zellenverbuende): **1496 us** je Abfrage gegenueber
    **7,5 us** fuer den Vergleich hier - Faktor 200.

    Das faellt auf, weil :func:`parse_grid` diese Abfrage rund 250.000 mal
    stellt: fuer jede Zelle, jede Fach- und jede Zustand-Zeile. Der alte Weg
    kostete damit gut drei Sekunden pro Aufruf, und der Dashboard-Server
    ruft ``parse_grid`` bei **jedem** Seitenaufruf und **jeder** Zellaenderung
    auf.

    Inhaltlich ist es dieselbe Pruefung: ``CellRange.__contains__`` vergleicht
    am Ende genau diese vier Grenzen. Deshalb kein Index und kein Cache - die
    Abfrage selbst ist billig genug, und ein Cache muesste ungueltig werden,
    sobald jemand einen Zellenverbund aendert.
    """
    return merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col


def resolve_anchor(ws, row: int, col: int) -> tuple[int, int]:
    """Gibt (anchor_row, anchor_col) zurueck - bei Zellenverbund die oben-links-Zelle."""
    for merged in ws.merged_cells.ranges:
        if merge_deckt_ab(merged, row, col):
            return merged.min_row, merged.min_col
    return row, col


def merged_range_at(ws, row: int, col: int):
    """Der Zellenverbund, der (row, col) ueberdeckt - oder None."""
    for merged in ws.merged_cells.ranges:
        if merge_deckt_ab(merged, row, col):
            return merged
    return None


def find_fach_for_col(ws, fach_rows: list[int], col: int) -> str | None:
    """Fach-Label fuer Spalte col aus der naechsten (untersten) Fach-Zeile mit Inhalt.

    Ist die Zelle leer, wird die naechsthoehere Fach-Zeile als Fallback genommen -
    im echten Blatt traegt die Oberstufen-Fachzeile nur die abweichenden Faecher.
    """
    for fach_row in reversed(fach_rows):
        ar, ac = resolve_anchor(ws, fach_row, col)
        if ar != fach_row:
            # Anker liegt in einer anderen Zeile - diese Zeile ueberspringen
            continue
        val = ws.cell(ar, ac).value
        if val is not None:
            return str(val)
    return None


def find_zustand_for_col(ws, zustand_rows: list[int], col: int) -> str | None:
    """Zustand-Label fuer Spalte col, mit Fallback auf hoehere Zustand-Zeilen."""
    for zustand_row in reversed(zustand_rows):
        ar, ac = resolve_anchor(ws, zustand_row, col)
        if ar != zustand_row:
            continue
        val = ws.cell(ar, ac).value
        if val is not None:
            return str(val)
    return None


# ── Zeilen-Klassifikation ─────────────────────────────────────────────────────

def classify_row(ws, row: int) -> str:
    """'fach' | 'zustand' | 'stand' | 'jahrgang' | 'other'"""
    val = ws.cell(row, 1).value
    if val == "Fach":
        return "fach"
    if val == "Zustand":
        return "zustand"
    if val == "Stand":
        return "stand"
    if isinstance(val, str) and re.match(r"Jahrgang\s+\d+", val):
        return "jahrgang"
    return "other"


def extract_grade(ws, row: int) -> int | None:
    val = ws.cell(row, 1).value
    m = re.match(r"Jahrgang\s+(\d+)", str(val)) if val else None
    return int(m.group(1)) if m else None


def strip_hint(text: str) -> tuple[str, str | None]:
    """Trennt Serientitel-Hinweis in Klammern vom Fach-Namen.

    'Politik (eA)' -> ('Politik', 'eA');  'Deutsch' -> ('Deutsch', None)
    """
    m = re.match(r"^(.*?)\s*\((.+)\)\s*$", text.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return text.strip(), None


# ── Datenmodell ───────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class GridCell:
    """Eine (Jahrgangszeile, Spalte)-Kreuzung, in Laufreihenfolge des Blatts.

    Auch nicht verwendbare Kreuzungen bleiben erhalten (``skip_reason`` gesetzt),
    damit Diagnosen und Debug-Ausgaben genau in Blattreihenfolge entstehen.
    """
    row: int
    col: int
    grade: int
    zustand_label: str | None
    zustand: str | None
    fach_label: str | None
    subject: str | None
    hint: str | None
    anchor_row: int
    anchor_col: int
    skip_reason: str | None

    @property
    def col_letter(self) -> str:
        return get_column_letter(self.col)

    @property
    def anchor_ref(self) -> str:
        return f"{get_column_letter(self.anchor_col)}{self.anchor_row}"

    @property
    def usable(self) -> bool:
        return self.skip_reason is None


@dataclass(frozen=True)
class GridSlot:
    """Eine Zustandszelle einer Listenzeile."""
    zustand: str
    ref: str
    row: int
    col: int
    span_rows: tuple[int, ...]


@dataclass(frozen=True)
class GridEntry:
    """Eine Zeile der spaeteren flachen Liste - ein Buch in einem Jahrgangsband."""
    key: str
    block: int
    fach_label: str
    subject: str
    hint: str | None
    grades: tuple[int, ...]
    slots: dict[str, GridSlot]
    angemeldet_refs: tuple[str, ...]

    @property
    def grade_label(self) -> str:
        if not self.grades:
            return ""
        if len(self.grades) == 1:
            return str(self.grades[0])
        return f"{self.grades[0]}-{self.grades[-1]}"


@dataclass(frozen=True)
class Grid:
    cells: tuple[GridCell, ...]
    entries: tuple[GridEntry, ...]
    stand_rows: tuple[int, ...]
    blocked: tuple[str, ...]
    fach_rows: tuple[int, ...]
    zustand_rows: tuple[int, ...]
    blocks: tuple[tuple[int, int], ...]

    def entry(self, key: str) -> GridEntry | None:
        for entry in self.entries:
            if entry.key == key:
                return entry
        return None


# ── Parser ────────────────────────────────────────────────────────────────────

def find_blocks(ws, zustand_rows: list[int], max_col: int) -> list[tuple[int, int]]:
    """Fachbloecke als (erste Spalte, letzte Spalte).

    Ein Block beginnt an jeder Spalte mit Zustand 'Angemeldet' und endet vor der
    naechsten. Das ist robuster als sich auf die Merges der Fach-Zeile zu
    verlassen, die ein Bearbeiter jederzeit aufloesen kann.
    """
    starts: list[int] = []
    last_labelled = 1
    for col in range(2, max_col + 1):
        label = find_zustand_for_col(ws, zustand_rows, col)
        if label is None:
            continue
        last_labelled = col
        if label.strip().lower() == BLOCK_START_ZUSTAND:
            starts.append(col)
    return [
        (start, (starts[i + 1] - 1) if i + 1 < len(starts) else last_labelled)
        for i, start in enumerate(starts)
    ]


def _blocked_refs(ws, blocks: list[tuple[int, int]], jahrgang_rows: set[int]) -> tuple[set[str], list[str]]:
    """Sperrflaechen: Merges ueber die volle Blockbreite in Jahrgangszeilen.

    Rueckgabe: (Menge aller ueberdeckten Zellreferenzen, sortierte Ankerliste).
    """
    covered: set[str] = set()
    anchors: list[str] = []
    widths = {start: end for start, end in blocks}
    for merged in ws.merged_cells.ranges:
        if widths.get(merged.min_col) != merged.max_col:
            continue
        if merged.min_col == merged.max_col:
            continue  # einspaltiger Block waere kein Fachblock
        rows = set(range(merged.min_row, merged.max_row + 1))
        if not rows & jahrgang_rows:
            continue
        anchors.append(merged.coord)
        for row in rows:
            for col in range(merged.min_col, merged.max_col + 1):
                covered.add(f"{get_column_letter(col)}{row}")
    return covered, sorted(anchors)


def parse_grid(ws, *, skip_blocked: bool = True, debug: Debug = None) -> Grid:
    """Liest die Struktur des Bestandsblatts. Kein Netz, kein Schreiben.

    ``skip_blocked=False`` schaltet Regel 4 ab und laesst Sperrflaechen wie
    gewoehnliche Zellen behandeln - nur fuer den Vergleich mit dem alten Verhalten.
    """
    def _dbg(msg: str) -> None:
        if debug is not None:
            debug(msg)

    # Erster Durchgang: Zeilenarten bestimmen. Die Sperrflaechen-Regel braucht
    # alle Jahrgangszeilen im Voraus, die Blockbreiten alle Zustand-Zeilen.
    row_types = {row: classify_row(ws, row) for row in range(1, ws.max_row + 1)}
    all_zustand_rows = [row for row, kind in row_types.items() if kind == "zustand"]
    all_jahrgang_rows = {
        row for row, kind in row_types.items()
        if kind == "jahrgang" and extract_grade(ws, row) is not None
    }
    blocks = find_blocks(ws, all_zustand_rows, ws.max_column)
    covered, blocked_anchors = _blocked_refs(ws, blocks, all_jahrgang_rows)

    # Zweiter Durchgang: von oben nach unten. Fach- und Zustand-Zeilen gelten erst
    # ab der Zeile, in der sie stehen - die Oberstufen-Fachzeile darf die
    # Sek-I-Zeilen darueber nicht umdeuten.
    fach_rows: list[int] = []
    zustand_rows: list[int] = []
    stand_rows: list[int] = []
    cells: list[GridCell] = []
    consecutive_other = 0

    for row in range(1, ws.max_row + 1):
        row_type = row_types[row]
        if row_type == "fach":
            fach_rows.append(row)
            consecutive_other = 0
            _dbg(f"  Zeile {row}: FACH erkannt")
            continue
        if row_type == "zustand":
            zustand_rows.append(row)
            consecutive_other = 0
            _dbg(f"  Zeile {row}: ZUSTAND erkannt")
            continue
        if row_type == "stand":
            stand_rows.append(row)
            consecutive_other = 0
            _dbg(f"  Zeile {row}: STAND erkannt")
            continue
        if row_type == "other":
            consecutive_other += 1
            _dbg(f"  Zeile {row}: other (consecutive={consecutive_other}, "
                 f"A={ws.cell(row, 1).value!r})")
            continue

        consecutive_other = 0
        grade = extract_grade(ws, row)
        _dbg(f"\n  Zeile {row}: JAHRGANG {grade} | fach_rows={fach_rows} | "
             f"zustand_rows={zustand_rows}")
        if grade is None or not fach_rows:
            _dbg(f"    -> Uebersprungen (grade={grade}, fach_rows leer={not fach_rows})")
            continue

        for col in range(2, ws.max_column + 1):
            zustand_label = find_zustand_for_col(ws, zustand_rows, col)
            zustand = zustand_label.strip().lower() if zustand_label else None
            fach_label = None
            subject = hint = None
            anchor_row, anchor_col = resolve_anchor(ws, row, col)

            if zustand_label is None:
                reason: str | None = SKIP_NO_ZUSTAND
            elif zustand not in WRITABLE_ZUSTAENDE:
                reason = SKIP_ZUSTAND_NOT_WRITABLE
            else:
                fach_label = find_fach_for_col(ws, fach_rows, col)
                if fach_label is None:
                    reason = SKIP_NO_FACH
                else:
                    subject, hint = strip_hint(fach_label)
                    ref = f"{get_column_letter(col)}{row}"
                    reason = SKIP_NOT_OFFERED if (skip_blocked and ref in covered) else None

            cells.append(GridCell(
                row=row, col=col, grade=grade,
                zustand_label=zustand_label,
                zustand=zustand if reason != SKIP_NO_ZUSTAND else None,
                fach_label=fach_label, subject=subject, hint=hint,
                anchor_row=anchor_row, anchor_col=anchor_col, skip_reason=reason,
            ))

    entries = _build_entries(ws, cells, blocks)
    return Grid(
        cells=tuple(cells),
        entries=entries,
        stand_rows=tuple(stand_rows),
        blocked=tuple(blocked_anchors),
        fach_rows=tuple(fach_rows),
        zustand_rows=tuple(zustand_rows),
        blocks=tuple(blocks),
    )


def _span_rows(ws, row: int, col: int) -> tuple[int, ...]:
    merged = merged_range_at(ws, row, col)
    if merged is None:
        return (row,)
    return tuple(range(merged.min_row, merged.max_row + 1))


def _build_entries(ws, cells: list[GridCell], blocks: list[tuple[int, int]]) -> tuple[GridEntry, ...]:
    """Fasst die Zellen zu Listenzeilen zusammen; die Bestand-Spalte gruppiert.

    Der Merge der Bestand-Zelle definiert das Jahrgangsband. Angemeldet bleibt je
    Jahrgang einzeln und wird als Referenzliste mitgefuehrt, damit die Summe
    spaeter aus den einzelnen Zellen kommt.
    """
    block_of_col: dict[int, int] = {}
    for index, (start, end) in enumerate(blocks):
        for col in range(start, end + 1):
            block_of_col[col] = index

    by_row_col = {(cell.row, cell.col): cell for cell in cells}
    entries: list[GridEntry] = []
    seen: set[str] = set()

    for cell in cells:
        if not cell.usable or cell.zustand != "bestand":
            continue
        block = block_of_col.get(cell.col)
        if block is None:
            continue
        bestand_ref = cell.anchor_ref
        key = f"{block}:{cell.fach_label}:{bestand_ref}"
        if key in seen:
            continue
        seen.add(key)

        span = _span_rows(ws, cell.row, cell.col)
        grades: list[int] = []
        angemeldet_refs: list[str] = []
        for row in span:
            partner = by_row_col.get((row, blocks[block][0]))
            if partner is None or partner.grade in grades:
                continue
            grades.append(partner.grade)
            if partner.anchor_ref not in angemeldet_refs:
                angemeldet_refs.append(partner.anchor_ref)

        slots: dict[str, GridSlot] = {}
        start = blocks[block][0]
        for offset, name in enumerate(("angemeldet", "bestand", "bestellt", "zu_bestellen")):
            col = start + offset
            if col > blocks[block][1]:
                continue
            anchor_row, anchor_col = resolve_anchor(ws, cell.row, col)
            slots[name] = GridSlot(
                zustand=name,
                ref=f"{get_column_letter(anchor_col)}{anchor_row}",
                row=anchor_row, col=anchor_col,
                span_rows=_span_rows(ws, cell.row, col),
            )

        entries.append(GridEntry(
            key=key, block=block, fach_label=cell.fach_label or "",
            subject=cell.subject or "", hint=cell.hint,
            grades=tuple(grades), slots=slots,
            angemeldet_refs=tuple(angemeldet_refs),
        ))
    return tuple(entries)
