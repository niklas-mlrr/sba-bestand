"""Snapshot auf das Workbook anwenden und das Blatt 'zu Bestellen' neu aufbauen.

Reine Funktionen: sie bekommen ein geoeffnetes Workbook, ein geparstes
:class:`~bestand.core.grid.Grid` und einen :class:`~bestand.core.iserv.Snapshot`.
Kein Netz, kein ``print``, kein Speichern - das gehoert dem Aufrufer.

Das Workbook wird immer mit ``data_only=False`` geladen. Die Spalte
'zu bestellen' enthaelt echte Formeln; wer sie mit ``data_only=True`` laedt,
liest zwischengespeicherte Werte und schreibt die Formeln beim Speichern platt.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Callable, Optional

from ausleihe.inventory_excel import match_book
from openpyxl.utils import get_column_letter, range_boundaries

from .config import BestandConfig
from .grid import SKIP_NO_FACH, SKIP_NO_ZUSTAND, Blatt, Grid, GridCell
from .iserv import Buch, Snapshot

# Wie ``Blatt`` (siehe grid.py): openpyxl liefert kein py.typed.
Mappe = Any  # openpyxl.workbook.workbook.Workbook

try:
    import isbnlib as _isbnlib

    def format_isbn(isbn: str) -> str:
        try:
            masked = _isbnlib.mask(isbn)
            return masked if masked else isbn
        except Exception:
            return isbn
except ImportError:  # pragma: no cover - isbnlib ist eine harte Abhaengigkeit
    def format_isbn(isbn: str) -> str:
        return isbn


# Excel-Anzeigeformat der "Stand"-Zelle: z.B. "Dienstag, 30.06.2026 17:25:05"
# (TTTT, TT.MM.JJJJ hh:mm:ss). Punkte/Komma/Leerzeichen sind escaped (literal),
# der Wert selbst ist ein echtes datetime -> Excel rendert es ueber dieses Format.
STAND_NUMBER_FORMAT = r"dddd\,\ dd\.mm\.yyyy\ hh:mm:ss"

# Abkuerzungen, die nicht literal im Buchtitel stehen, sondern als Langform.
# Schluessel lowercase fuer case-insensitiven Lookup.
HINT_EXPANSIONS: dict[str, str] = {
    "ea": "Erhöhtes",
    "ga": "Grundlegendes",
}

Debug = Optional[Callable[[str], None]]


@dataclass(frozen=True)
class CellChange:
    ref: str
    old: Any
    new: Any
    note: str
    new_display: str | None = None

    def render(self) -> str:
        shown = self.new_display if self.new_display is not None else repr(self.new)
        return f"  {self.ref}: {self.old!r} -> {shown}  [{self.note}]"


@dataclass(frozen=True)
class ZuBestellenRow:
    grade: int
    fach: str
    stueckzahl: int
    title: str
    publisher: str
    isbn: str
    price: float


@dataclass
class UpdateResult:
    changes: list[CellChange] = field(default_factory=list)
    skipped: list[str] = field(default_factory=list)
    diagnostics: list[str] = field(default_factory=list)
    zu_bestellen_rows: list[ZuBestellenRow] = field(default_factory=list)
    zu_bestellen_data: dict[str, dict] = field(default_factory=dict)
    isbn_by_entry: dict[str, str] = field(default_factory=dict)
    stand: datetime | None = None

    @property
    def ok(self) -> bool:
        return not self.diagnostics


def load_bestellt_counts(ws_bestellt: Blatt) -> tuple[dict[str, int], list[str]]:
    """Liest Blatt 'bestellt': Spalte F = ISBN (evtl. mit '-'), Spalte C = Stueckzahl.

    Gibt pro normierter ISBN (ohne '-') die Summe aller Stueckzahlen zurueck.
    Zeile 1 (Kopfzeile) wird uebersprungen.
    """
    counts: dict[str, int] = {}
    errors: list[str] = []
    for row in range(2, ws_bestellt.max_row + 1):
        isbn_raw = ws_bestellt.cell(row, 6).value   # Spalte F
        count_raw = ws_bestellt.cell(row, 3).value  # Spalte C
        if isbn_raw is None and count_raw is None:
            continue
        if isbn_raw is None or count_raw is None:
            errors.append(f"bestellt!{row}: ISBN und Stückzahl müssen beide gesetzt sein.")
            continue
        isbn_norm = str(isbn_raw).replace("-", "").strip()
        if not isbn_norm:
            errors.append(f"bestellt!{row}: ISBN ist leer.")
            continue
        try:
            count = int(count_raw)
        except (ValueError, TypeError):
            errors.append(f"bestellt!{row}: ungültige Stückzahl {count_raw!r}.")
            continue
        counts[isbn_norm] = counts.get(isbn_norm, 0) + count
    return counts, errors


def _ganzzahl(wert: Any) -> int | None:
    """Ein vorhandener Zellwert als ganze Zahl - alles andere zaehlt als leer.

    Gebraucht fuer die "Bestellt"-Zelle, die :func:`apply_snapshot` stehen
    laesst: ihr Wert geht danach in ``zu_bestellen_data`` ein, und
    :func:`compute_zu_bestellen_rows` rechnet dort mit ``or 0``. Ein Text oder
    ein Formelrest wuerde diese Rechnung mit einem TypeError anhalten, ein
    ``bool`` waere stillschweigend 0 oder 1 - beides faengt diese Funktion ab.
    """
    if wert is None or isinstance(wert, (str, bool)):
        return None
    try:
        return int(wert)
    except (TypeError, ValueError):
        return None


def _entry_key(grid: Grid, cell: GridCell) -> str | None:
    for index, (start, end) in enumerate(grid.blocks):
        if start <= cell.col <= end:
            return f"{index}:{cell.fach_label}:{cell.anchor_ref}"
    return None


def apply_snapshot(
    ws: Blatt,
    grid: Grid,
    snapshot: Snapshot,
    config: BestandConfig,
    *,
    bestellt_counts: dict[str, int] | None = None,
    result: UpdateResult | None = None,
    stand: datetime | None = None,
    debug: Debug = None,
) -> UpdateResult:
    """Traegt Angemeldet/Bestand/Bestellt in die Ankerzellen ein.

    ``result`` kann eine vorbefuellte Instanz sein (z.B. mit den Diagnosen aus
    :func:`load_bestellt_counts`), damit deren Reihenfolge erhalten bleibt.
    Bei nicht leeren ``diagnostics`` darf der Aufrufer **nicht** speichern.

    Drei der vier Spalten kommen aus IServ (Angemeldet, Bezahlt, Bestand) und
    werden bedingungslos ueberschrieben. **Bestellt nicht**: es kommt aus dem
    Blatt ``bestellt`` derselben Mappe und wird nur dort geschrieben, wo die
    ISBN in ``bestellt_counts`` steht. Fehlt sie, bleibt die Zelle unberuehrt -
    Begruendung am Zweig selbst.
    """
    def _dbg(msg: str) -> None:
        if debug is not None:
            debug(msg)

    res = result if result is not None else UpdateResult()
    res.stand = stand or snapshot.fetched_at
    counts = bestellt_counts if bestellt_counts is not None else {}
    bestand_counts = snapshot.bestand_by_isbn

    processed_anchors: set[str] = set()                       # Zellenverbund-Dedup
    processed_bestand_isbns: set[str] = set()                 # Bestand: je ISBN einmal
    processed_bestellt_isbns: set[str] = set()                # Bestellt: je ISBN einmal
    processed_enrollment: set[tuple[int, str, str]] = set()   # (Jahrgang, ISBN, Zustand)

    current_row: int | None = None
    books: list[Buch] = []

    for cell in grid.cells:
        if cell.row != current_row:
            current_row = cell.row
            books = snapshot.books_for_grade(cell.grade)
            _dbg(f"    {len(books)} Bücher geladen")
            # Nicht `book`: so heisst unten das *ausgewaehlte* Buch, und das
            # darf None sein. Ein Name fuer zwei Dinge hiess, dass mypy den
            # nicht-optionalen Typ dieser Schleife festschrieb und die
            # Zuweisung `book = match.book` dagegen lief.
            for geladenes in books:
                _dbg(f"      isbn={geladenes['isbn']} subjects={geladenes['subjects']} "
                     f"title={geladenes['title'][:50]!r}")

        letter = cell.col_letter
        if cell.skip_reason == SKIP_NO_ZUSTAND:
            _dbg(f"    Sp.{letter}: kein Zustand-Label → skip")
            continue
        if cell.skip_reason and cell.zustand not in ("angemeldet", "bezahlt", "bestand", "bestellt"):
            _dbg(f"    Sp.{letter}: Zustand={cell.zustand_label!r} "
                 f"(nicht angemeldet/bezahlt/bestand/bestellt) → skip")
            continue
        # `subject is None` und SKIP_NO_FACH sind dieselbe Lage: parse_grid
        # setzt subject genau dann, wenn es ein Fach-Label gab (grid.py, ueber
        # strip_hint). Die Kopplung zweier Felder sieht mypy nicht, und seit
        # ausleihe-api py.typed liefert, verlangt match_book unten ein `str`.
        # Also wird die Bedingung ausgeschrieben statt nur der Grund abgefragt.
        if cell.skip_reason == SKIP_NO_FACH or cell.subject is None:
            res.diagnostics.append(
                f"Sp.{letter}/Zeile {cell.row}: kein Fach-Label für Zustand {cell.zustand_label!r}."
            )
            _dbg(f"    Sp.{letter}: [{cell.zustand_label}] kein Fach-Label → skip")
            continue
        if cell.skip_reason is not None:
            # Sperrflaeche: Merge ueber die volle Fachblockbreite = nicht angeboten.
            _dbg(f"    Sp.{letter}: [{cell.zustand_label}] Fach={cell.fach_label!r} "
                 f"→ {cell.skip_reason} → skip")
            continue

        override_key = f"{cell.grade}|{cell.subject}|{cell.hint or ''}"
        match = match_book(
            books, cell.subject, cell.hint,
            override_isbn=config.match_overrides.get(override_key),
            hint_expansions=HINT_EXPANSIONS,
        )
        book = match.book
        if book is None:
            if match.error and match.error.startswith("Kein Buch-Match"):
                # Fach in diesem Jahrgang schlicht nicht in der Buecherliste -
                # normal, kein Abbruch.
                res.skipped.append(f"Sp.{letter}/Zeile {cell.row}: {match.error}")
                _dbg(f"    Sp.{letter}: [{cell.zustand_label}] Fach={cell.fach_label!r} "
                     f"→ kein Buch-Match, übersprungen "
                     f"(subjects in Bücherliste: {[b['subjects'] for b in books]})")
            else:
                res.diagnostics.append(f"Sp.{letter}/Zeile {cell.row}: {match.error}")
                _dbg(f"    Sp.{letter}: [{cell.zustand_label}] Fach={cell.fach_label!r} "
                     f"→ {match.error}")
            continue

        anchor_ref = cell.anchor_ref
        if anchor_ref in processed_anchors:
            _dbg(f"    Sp.{letter}: {anchor_ref} bereits verarbeitet (Zellenverbund) → skip")
            continue
        processed_anchors.add(anchor_ref)

        isbn = book["isbn"]
        zustand = cell.zustand
        if zustand is None:
            # Unerreichbar, solange parse_grid ``zustand=None`` nur zusammen mit
            # einem ``skip_reason`` setzt (SKIP_NO_ZUSTAND beim fehlenden Label,
            # SKIP_ZUSTAND_NOT_WRITABLE beim leeren) - beide Faelle haben die
            # Waechter oben schon aussortiert. Steht hier, weil die Invariante
            # zwei Module ueberspannt und ``processed_enrollment`` sie annimmt:
            # ohne diese Zeile waere sein Schluessel
            # ``tuple[int, str, str | None]``, und ein parse_grid, das den
            # Zustand irgendwann ohne skip_reason leer laesst, zaehlte hier
            # still eine None-Zeile mit statt aufzufallen.
            _dbg(f"    Sp.{letter}: kein Zustand trotz skip_reason=None → skip")
            continue
        if zustand == "bestand":
            if isbn in processed_bestand_isbns:
                _dbg(f"    Sp.{letter}: isbn={isbn}/Bestand bereits eingetragen → skip")
                continue
            processed_bestand_isbns.add(isbn)
            key = _entry_key(grid, cell)
            if key is not None:
                res.isbn_by_entry[key] = isbn
        elif zustand == "bestellt":
            if isbn in processed_bestellt_isbns:
                _dbg(f"    Sp.{letter}: isbn={isbn}/Bestellt bereits eingetragen → skip")
                continue
            processed_bestellt_isbns.add(isbn)
        else:
            enr_key = (cell.grade, isbn, zustand)
            if enr_key in processed_enrollment:
                _dbg(f"    Sp.{letter}: isbn={isbn}/{cell.zustand_label} Jg.{cell.grade} "
                     f"bereits eingetragen → skip")
                continue
            processed_enrollment.add(enr_key)

        lookup = (cell.grade, isbn)
        # "Bestellt" ist die einzige Spalte, die auch von Hand gepflegt wird -
        # deshalb kann sie hier ausgelassen werden, siehe unten.
        behalten = False
        if zustand == "angemeldet":
            new_val: int | None = snapshot.enrolled.get(lookup, 0)
        elif zustand == "bezahlt":
            new_val = snapshot.paid.get(lookup, 0)
        elif zustand == "bestand":
            new_val = bestand_counts.get(isbn, 0)
        else:  # bestellt
            isbn_norm = isbn.replace("-", "")
            if isbn_norm in counts:
                new_val = counts[isbn_norm]
            else:
                # Bis 2026-09-05 stand hier None, die Zelle wurde also GELEERT.
                # Das war der einzige Ort, an dem ein Abruf eine Eingabe von
                # Hand vernichtete: "Bestellt" kommt nicht aus IServ, sondern
                # aus dem Blatt 'bestellt' derselben Mappe, und wer eine
                # Bestellung dort (noch) nicht eingetragen hat, aber die Zahl
                # schon im Raster stehen hatte, fand sie nach dem Abruf leer
                # vor. Ein fehlender Eintrag im Blatt heisst "ich weiss nichts
                # darueber", nicht "es ist nichts bestellt" - deshalb bleibt
                # die Zelle jetzt stehen, und der vorhandene Wert geht
                # unveraendert in die Bedarfsrechnung ein.
                behalten = True
                new_val = _ganzzahl(ws[anchor_ref].value)

        entry = res.zu_bestellen_data.setdefault(isbn, {
            "angemeldet": 0,
            "bestand": 0,
            "bestellt": None,
            "title": book["title"],
            "grades": set(),
            "fach": cell.fach_label,
        })
        entry["grades"].add(cell.grade)
        if zustand == "angemeldet" and new_val is not None:
            entry["angemeldet"] = (entry["angemeldet"] or 0) + new_val
        elif zustand == "bestand" and new_val is not None:
            entry["bestand"] = new_val
        elif zustand == "bestellt":
            entry["bestellt"] = new_val

        if debug is not None:
            isbn_norm = isbn.replace("-", "")
            _dbg(
                f"    Sp.{letter}: {anchor_ref} {cell.fach_label!r}/{cell.zustand_label} → "
                f"isbn={isbn}, enrolled={snapshot.enrolled.get(lookup, '–')}, "
                f"paid={snapshot.paid.get(lookup, '–')}, bestand={bestand_counts.get(isbn, '–')}, "
                f"bestellt_sheet={counts.get(isbn_norm, '–')}, new_val={new_val}"
            )

        if behalten:
            # Nichts schreiben und nichts melden: eine Zelle, die der Abruf
            # bewusst nicht anfasst, ist keine Aenderung. Stuende sie trotzdem
            # in res.changes, meldete die Zusammenfassung sie als "aktualisiert"
            # und das Dashboard hoebe sie hinterher als frisch geaendert hervor.
            continue

        old_val = ws[anchor_ref].value
        ws[anchor_ref].value = new_val

        hint_str = f" ({cell.hint})" if cell.hint else ""
        res.changes.append(CellChange(
            ref=anchor_ref, old=old_val, new=new_val,
            note=f"{cell.subject}{hint_str} Jg.{cell.grade}, {book['title']}, {cell.zustand_label}",
        ))

    return res


def write_stand(
    ws: Blatt, grid: Grid, when: datetime | None, result: UpdateResult
) -> UpdateResult:
    """Traegt den Abfragezeitpunkt in Spalte B jeder erkannten 'Stand'-Zeile ein.

    ``when=None`` heisst "nimm ``result.stand``" - den Zeitpunkt, den
    :func:`apply_snapshot` dort in jedem Fall hinterlaesst
    (``res.stand = stand or snapshot.fetched_at``). Der Parameter nimmt
    ``None`` an, weil beide Aufrufer ``write_stand(ws, grid, result.stand,
    result)`` schreiben und ``UpdateResult.stand`` als Feld mit Default
    ``datetime | None`` sein *muss*: eine frisch gebaute ``UpdateResult``
    (das Dashboard baut eine, um die Diagnosen von ``load_bestellt_counts``
    hineinzureichen) hat noch keinen Stand. Ohne diese Zeile meldete mypy den
    Aufruf an genau drei Stellen - hier, in ``update_bestand_auto.py`` und in
    ``sba-dashboard/app/refresh.py`` - als Typfehler, obwohl keine davon je
    ``None`` durchreicht.
    """
    from .grid import resolve_anchor

    zeitpunkt = when if when is not None else result.stand
    if zeitpunkt is None:
        # Erreichbar nur bei Fehlgebrauch: write_stand mit einer UpdateResult,
        # die nie durch apply_snapshot lief. Lieber hier laut, als eine leere
        # Stand-Zelle in die Mappe zu schreiben.
        raise ValueError(
            "write_stand braucht einen Zeitpunkt: 'when' ist None und "
            "result.stand ist leer - lief diese UpdateResult durch apply_snapshot?"
        )

    for stand_row in grid.stand_rows:
        anchor_row, anchor_col = resolve_anchor(ws, stand_row, 2)  # Spalte B
        cell = ws.cell(anchor_row, anchor_col)
        ref = f"{get_column_letter(anchor_col)}{anchor_row}"
        old_val = cell.value
        cell.value = zeitpunkt                   # echtes datetime (Excel-Datumswert)
        cell.number_format = STAND_NUMBER_FORMAT  # Anzeige: TTTT, TT.MM.JJJJ hh:mm:ss
        result.changes.append(CellChange(
            ref=ref, old=old_val, new=zeitpunkt, note="Stand/Abfragezeitpunkt",
            new_display=zeitpunkt.strftime("%d.%m.%Y %H:%M:%S"),
        ))
    return result


# ── Blatt "zu Bestellen" ──────────────────────────────────────────────────────

_SUBTOTAL_FUNC = {
    "sum": 109, "count": 103, "countNums": 102, "average": 101,
    "max": 104, "min": 105, "stdDev": 107, "var": 110,
}


def _norm_isbn(value: Any) -> str:
    return re.sub(r"[^0-9Xx]", "", str(value)) if value is not None else ""


def compute_zu_bestellen_rows(
    zu_bestellen_data: dict[str, dict], series_data: dict[str, dict], safety_stock: int
) -> list[ZuBestellenRow]:
    """Bedarf = Angemeldet - Bestand - Bestellt, in Python gerechnet.

    Nie aus der Formel der Spalte 'zu bestellen' gelesen: die steht als Text im
    Workbook und haette ohne Excel-Neuberechnung keinen Wert.
    """
    rows: list[ZuBestellenRow] = []
    for isbn, entry in zu_bestellen_data.items():
        bedarf = (entry["angemeldet"] or 0) - (entry["bestand"] or 0) - (entry["bestellt"] or 0)
        if bedarf <= 0:
            continue
        sd = series_data.get(isbn, {})
        rows.append(ZuBestellenRow(
            grade=min(entry["grades"]),
            fach=entry["fach"],
            stueckzahl=bedarf + safety_stock,
            title=sd.get("title") or entry["title"],
            publisher=sd.get("publisher", ""),
            isbn=format_isbn(isbn),
            price=sd.get("price", 0.0),
        ))
    rows.sort(key=lambda r: r.title)
    return rows


def rebuild_zu_bestellen(
    wb: Mappe, result: UpdateResult, snapshot: Snapshot, safety_stock: int
) -> list[ZuBestellenRow]:
    """Baut das Blatt 'zu Bestellen' neu auf und passt die Tabellengeometrie an.

    Der Tabellenname wird dynamisch gelesen - er kann sich aendern. Vorhandene
    'Bestell Nr.' wandern per ISBN wieder an ihre Buchreihe.
    """
    if "zu Bestellen" not in wb.sheetnames:
        raise KeyError("Blatt 'zu Bestellen' fehlt.")
    ws_zu = wb["zu Bestellen"]
    if not ws_zu.tables:
        raise RuntimeError("Blatt 'zu Bestellen' enthält keine Tabelle.")

    table_name = next(iter(ws_zu.tables))
    table = ws_zu.tables[table_name]
    if len(table.tableColumns) < 8:
        raise ValueError("Tabelle auf 'zu Bestellen' hat nicht die erwarteten acht Eingabespalten.")

    t_min_col, header_row, t_max_col, t_old_max_row = range_boundaries(table.ref)
    totals_count = table.totalsRowCount or 0
    first_data_row = header_row + 1
    old_data_last = t_old_max_row - totals_count
    bestellnr_col = t_min_col    # Spalte A: "Bestell Nr."
    gesamtpreis_col = t_max_col  # Spalte I: "Gesamtpreis (brutto)" (Formel)

    qty_col_name = table.tableColumns[3].name    # "Stückzahl"
    price_col_name = table.tableColumns[7].name  # "Einzelpreis (brutto)"

    def gesamtpreis_formula() -> str:
        return (
            f'=IF(OR({table_name}[[#This Row],[{qty_col_name}]]="",'
            f'{table_name}[[#This Row],[{price_col_name}]]=""),"",'
            f'{table_name}[[#This Row],[{qty_col_name}]]*'
            f'{table_name}[[#This Row],[{price_col_name}]])'
        )

    isbn_col = 7  # Spalte G
    bestellnr_by_isbn: dict[str, object] = {}
    for row in range(first_data_row, old_data_last + 1):
        isbn_key = _norm_isbn(ws_zu.cell(row, isbn_col).value)
        a_val = ws_zu.cell(row, bestellnr_col).value
        if isbn_key and a_val is not None:
            bestellnr_by_isbn[isbn_key] = a_val

    rows = compute_zu_bestellen_rows(
        result.zu_bestellen_data, snapshot.series_data, safety_stock
    )
    result.zu_bestellen_rows = rows

    # Gesamten vorherigen Tabellenbereich leeren, damit beim Verkleinern
    # nichts aus dem alten Bereich stehen bleibt.
    for row in range(first_data_row, t_old_max_row + 1):
        for col in range(t_min_col, t_max_col + 1):
            ws_zu.cell(row, col).value = None

    for i, entry in enumerate(rows):
        row = first_data_row + i
        ws_zu.cell(row, bestellnr_col).value = bestellnr_by_isbn.get(_norm_isbn(entry.isbn))
        ws_zu.cell(row, 2).value = entry.grade
        ws_zu.cell(row, 3).value = entry.fach
        ws_zu.cell(row, 4).value = entry.stueckzahl
        ws_zu.cell(row, 5).value = entry.title
        ws_zu.cell(row, 6).value = entry.publisher
        ws_zu.cell(row, 7).value = entry.isbn
        ws_zu.cell(row, 8).value = entry.price

    # Excel verlangt mindestens eine Datenzeile -> bei 0 Treffern eine leere.
    n_data = max(1, len(rows))
    data_last_row = header_row + n_data
    new_last_row = data_last_row + totals_count

    for row in range(first_data_row, data_last_row + 1):
        ws_zu.cell(row, gesamtpreis_col).value = gesamtpreis_formula()

    if totals_count:
        totals_row = new_last_row
        for offset, tc in enumerate(table.tableColumns):
            cell = ws_zu.cell(totals_row, t_min_col + offset)
            if tc.totalsRowLabel:
                cell.value = tc.totalsRowLabel
            elif tc.totalsRowFunction and tc.totalsRowFunction in _SUBTOTAL_FUNC:
                cell.value = f"=SUBTOTAL({_SUBTOTAL_FUNC[tc.totalsRowFunction]},{table_name}[{tc.name}])"
            else:
                cell.value = None

    def _ref(last_row: int) -> str:
        return (
            f"{get_column_letter(t_min_col)}{header_row}:"
            f"{get_column_letter(t_max_col)}{last_row}"
        )

    table.ref = _ref(new_last_row)
    if table.autoFilter is not None:
        table.autoFilter.ref = _ref(data_last_row)
    table.sortState = None  # alten Sortierbereich verwerfen (sonst ungueltig)
    return rows
