#!/usr/bin/env python3
"""Auto-Discovery: befüllt 'Angemeldet'-, 'Bestand'- und 'Bestellt'-Zellen anhand
der Excel-Struktur und trägt den Nachbestellbedarf in das Blatt 'zu Bestellen' ein.

Dieses Skript ist nur noch die Kommandozeilen-Schale. Die Logik liegt in
``bestand/core/`` und ist ohne Netz und ohne echte Excel-Datei testbar:

    core/config.py  config.json laden und prüfen
    core/grid.py    Excel-Struktur lesen (Fachblöcke, Mehrjahresbände, Sperrflächen)
    core/iserv.py   Snapshot aus IServ holen (nur GET)
    core/update.py  Snapshot anwenden, Blatt 'zu Bestellen' neu aufbauen

Verwendung:
    python3 update_bestand_auto.py [--dry-run] [--schoolyear 2025/2026]
                                   [--excel "Bestand- und Nachbestellungsliste 2026.xlsx"]
                                   [--sheet "Bestand- und Nachbestellung"]
                                   [--safety-stock 5] [--no-backup] [-v | --verbose]

Nur GET-Zugriffe auf die API. Kein Schreiben in die Datenbank.
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

_HERE = Path(__file__).parent
_ROOT = _HERE.parent

# sba-bestand hält keine eigenen Secrets: Die IServ-Credentials liegen in der
# ``.env`` des Geschwister-Repos ausleihe-api. Beide Repos werden nebeneinander
# geklont (``<irgendein-ordner>/{ausleihe-api,sba-bestand}``).
_API_ROOT = _ROOT.parent / "ausleihe-api"

# ``ausleihe`` kommt normalerweise aus dem Venv (editable-Install, siehe
# pyproject). Fallback auf das Geschwister-Repo, damit das Skript auch ohne
# Installation läuft (Nachfolger-Pfad: nur klonen, nichts installieren).
if _API_ROOT.is_dir():
    sys.path.insert(0, str(_API_ROOT))
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from dotenv import load_dotenv  # noqa: E402

load_dotenv(_API_ROOT / ".env")

from ausleihe import AusleiheClient  # noqa: E402
from ausleihe.inventory_excel import atomic_save_workbook  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from bestand.core import (  # noqa: E402
    EV_BOOKLISTS,
    EV_ENROLLMENTS,
    EV_GRADE_BOOKS,
    EV_NO_BOOKLIST,
    EV_SERIES,
    BestandConfig,
    ConfigError,
    UpdateResult,
    apply_snapshot,
    fetch_snapshot,
    load_bestellt_counts,
    parse_grid,
    rebuild_zu_bestellen,
    write_stand,
)

# Regel 4 (Sperrflächen) ist bewusst abschaltbar, damit sich das Verhalten vor
# und nach ihrer Einführung vergleichen lässt.
SKIP_BLOCKED = True


def _progress_printer(sy_id: str):
    """Bildet die Fortschritts-Ereignisse des Kerns auf die bisherigen Zeilen ab."""
    texte = {
        EV_BOOKLISTS: lambda p: "Lade Bücherlisten...",
        EV_ENROLLMENTS: lambda p: "Lade Anmeldungen...",
        EV_SERIES: lambda p: "Lade Serien-Daten...",
        EV_GRADE_BOOKS: lambda p: f"  Lade Bücherliste Jahrgang {p['grade']}...",
        EV_NO_BOOKLIST: lambda p: (
            f"  WARNUNG: Keine Bücherliste für Jahrgang {p['grade']} im Schuljahr {sy_id}"
        ),
    }

    def emit(event: str, payload: dict) -> None:
        render = texte.get(event)
        if render is not None:
            print(render(payload))

    return emit


def main(argv: list[str] | None = None, *, client_factory=AusleiheClient) -> None:
    parser = argparse.ArgumentParser(
        description="Auto-Befüllung der Bestandsliste anhand der Excel-Struktur"
    )
    parser.add_argument("--dry-run", action="store_true", help="Keine Änderungen speichern")
    parser.add_argument("--schoolyear", help="Schuljahr-ID, z.B. 2025/2026")
    parser.add_argument("--excel", help="Excel-Dateiname (überschreibt config.json)")
    parser.add_argument("--sheet", help="Tabellenblatt-Name (überschreibt config.json)")
    parser.add_argument("--safety-stock", type=int, help="Zusätzlicher Sicherheitsbestand je Titel")
    parser.add_argument("--no-backup", action="store_true", help="Kein Wiederherstellungs-Backup anlegen")
    parser.add_argument("-v", "--verbose", action="store_true", help="Detaillierte Debug-Ausgaben")
    args = parser.parse_args(argv)

    dbg = (lambda msg: print(msg)) if args.verbose else None

    print(f"Verbinde mit IServ ({os.environ.get('ISERV_DOMAIN', '?')})...")
    client = client_factory()

    sy_id = args.schoolyear or client.schoolyears.get_current()["id"]
    print(f"Schuljahr: {sy_id}")

    try:
        config = BestandConfig.load(
            _HERE / "config.json",
            excel=args.excel,
            sheet=args.sheet,
            safety_stock=args.safety_stock,
        )
    except ConfigError as exc:
        raise SystemExit(str(exc))
    print(f"Excel: {config.excel_path.name}  |  Blatt: {config.sheet_name}")

    snapshot = fetch_snapshot(client, sy_id, progress=_progress_printer(sy_id))

    if args.verbose:
        print(f"  enrolled_counts: {len(snapshot.enrolled)} Einträge, "
              f"paid_counts: {len(snapshot.paid)} Einträge")
        print(f"  series_data:     {len(snapshot.series_data)} Serien")
        for key, value in list(snapshot.enrolled.items())[:5]:
            sd = snapshot.series_data.get(key[1], {})
            print(f"    {key}: enrolled={value}, paid={snapshot.paid.get(key, 0)}, "
                  f"bestand={sd.get('total', 0)}")

    # data_only=False: 'zu bestellen' sind echte Formeln und müssen es bleiben.
    wb = load_workbook(str(config.excel_path))
    if config.sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet {config.sheet_name!r} fehlt.")
    ws = wb[config.sheet_name]

    if "bestellt" not in wb.sheetnames:
        raise SystemExit("Sheet 'bestellt' fehlt.")
    bestellt_counts, diagnostics = load_bestellt_counts(wb["bestellt"])
    if args.verbose:
        print(f"\nbestellt-Sheet: {len(bestellt_counts)} ISBNs mit Bestellungen")
        for isbn_norm, count in bestellt_counts.items():
            print(f"  {isbn_norm}: {count}")

        print(f"\nExcel geladen: {ws.max_row} Zeilen, {ws.max_column} Spalten")
        print("Erste Zeilen (Spalte A):")
        for r in range(1, min(ws.max_row + 1, 20)):
            print(f"  Zeile {r:3d}: A={ws.cell(r, 1).value!r}")

    print("\nAnalysiere Excel-Struktur...\n")

    grid = parse_grid(ws, skip_blocked=SKIP_BLOCKED, debug=dbg)
    result = apply_snapshot(
        ws, grid, snapshot, config,
        bestellt_counts=bestellt_counts,
        result=UpdateResult(diagnostics=list(diagnostics)),
        debug=dbg,
    )

    if result.skipped:
        print(f"\n{len(result.skipped)} Fach/Jahrgang-Kombination(en) ohne Buch übersprungen:")
        for line in result.skipped:
            print(f"  - {line}")

    if result.diagnostics:
        print("\nABBRUCH: Excel-Struktur oder Buch-Zuordnung ist nicht eindeutig; "
              "keine Datei wird gespeichert.")
        for line in result.diagnostics:
            print(f"  - {line}")
        raise SystemExit(2)

    write_stand(ws, grid, result.stand, result)

    try:
        rows = rebuild_zu_bestellen(wb, result, snapshot, config.safety_stock)
    except (KeyError, ValueError) as exc:
        raise SystemExit(str(exc))

    print(f"\n{len(rows)} Bücher mit Nachbestellbedarf:")
    for row in rows:
        print(f"  Jg.{row.grade:2d} [{row.fach}] +{config.safety_stock} → "
              f"{row.stueckzahl} Stk.  {row.title[:45]}  [{row.isbn}]")

    print()
    if args.dry_run:
        print("-- DRY RUN: keine Datei wird gespeichert --")

    if result.changes:
        verb = "würden aktualisiert" if args.dry_run else "aktualisiert"
        print(f"{len(result.changes)} Zelle(n) {verb}:")
        for change in result.changes:
            print(change.render())
    else:
        print("Keine Änderungen.")

    if not args.dry_run:
        backup_dir = None if args.no_backup else config.excel_path.parent / "backups"
        backup_path = atomic_save_workbook(wb, config.excel_path, backup_dir=backup_dir)
        print(f"\nGespeichert: {config.excel_path}")
        if backup_path:
            print(f"Backup: {backup_path}")


if __name__ == "__main__":
    main()
