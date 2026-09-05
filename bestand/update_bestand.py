#!/usr/bin/env python3
"""Aktualisiert die 'Bestand'- und 'Angemeldet'-Zellen in der Excel-Datei mit Daten aus der Ausleihe-API.

Verwendung:
    python3 update_bestand.py [--dry-run] [--schoolyear 2025/2026]

Liest config.json aus demselben Verzeichnis. Alle Pfade sind relativ zur
Position dieser Datei – das Skript ist vollständig selbstständig.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

_HERE = Path(__file__).parent
_ROOT = _HERE.parent

# sba-bestand hält keine eigenen Secrets: Die IServ-Credentials liegen in der
# ``.env`` des Geschwister-Repos ausleihe-api. Beide Repos werden nebeneinander
# geklont (``<irgendein-ordner>/{ausleihe-api,sba-bestand}``).
_API_ROOT = _ROOT.parent / "ausleihe-api"

# ``ausleihe`` kommt normalerweise aus dem Venv (editable-Install, siehe
# pyproject). Fallback auf das Geschwister-Repo, damit das Skript auch ohne
# Installation läuft (Nachfolger-Pfad: nur klonen, nichts installieren).
if _API_ROOT.is_dir():
    sys.path.insert(0, str(_API_ROOT))
# Beim Direktaufruf liegt nur ``bestand/`` auf dem Pfad, nicht die Repo-Wurzel;
# ``bestand.core`` waere sonst nicht importierbar.
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from dotenv import load_dotenv

load_dotenv(_API_ROOT / ".env")

from ausleihe import AusleiheClient, NotFoundError
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from bestand.core import atomic_save_workbook


def resolve_anchor(ws: Any, cell_ref: str) -> str:
    """Gibt die Ankerzelle (oben-links) zurück, falls cell_ref Teil einer merged-range ist."""
    for merged in ws.merged_cells.ranges:
        if cell_ref in merged:
            return f"{get_column_letter(merged.min_col)}{merged.min_row}"
    return cell_ref


def load_config() -> dict:
    path = _HERE / "config.json"
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def pick_schoolyear(client: AusleiheClient, override: str | None) -> str:
    """Wählt das Schuljahr: explizit per --schoolyear oder das aktuelle (heute zwischen begin/end, sonst neuestes nicht-archiviertes)."""
    years = client.admin.get_schoolyears()
    if override:
        ids = [y["id"] for y in years]
        if override not in ids:
            raise SystemExit(f"Schuljahr {override!r} nicht gefunden. Verfügbar: {ids}")
        return override
    now = datetime.now(timezone.utc)
    for y in years:
        if y.get("archived_at"):
            continue
        begin = datetime.fromisoformat(y["begin"].replace("Z", "+00:00"))
        end = datetime.fromisoformat(y["end"].replace("Z", "+00:00"))
        if begin <= now <= end:
            return y["id"]
    candidates = [y for y in years if not y.get("archived_at")]
    if not candidates:
        raise SystemExit("Kein passendes Schuljahr gefunden.")
    return candidates[-1]["id"]


def fetch_enrollment_counts(client: AusleiheClient, schoolyear_id: str) -> dict[str, int]:
    """Zählt aktive Anmeldungen pro ISBN für das angegebene Schuljahr."""
    enrollments = client.admin.get_enrollments(schoolyear_id)
    counts: dict[str, int] = {}
    for enr in enrollments:
        if enr.get("deleted_at"):
            continue
        isbns_in_enrollment = {
            item["series"]
            for item in enr.get("booklistItems", [])
            if item.get("series")
        }
        for isbn in isbns_in_enrollment:
            counts[isbn] = counts.get(isbn, 0) + 1
    return counts


def main() -> None:
    parser = argparse.ArgumentParser(description="Bestand- und Angemeldet-Zellen in Excel aktualisieren")
    parser.add_argument("--dry-run", action="store_true", help="Keine Änderungen speichern")
    parser.add_argument("--schoolyear", help="Schuljahr-ID, z.B. 2025/2026 (Default: aktuelles)")
    parser.add_argument("--allow-partial", action="store_true", help="Trotz nicht erreichbarer Serien speichern (nicht empfohlen)")
    parser.add_argument("--no-backup", action="store_true", help="Kein Wiederherstellungs-Backup anlegen")
    args = parser.parse_args()

    print("WARNUNG: update_bestand.py ist veraltet. Bitte update_bestand_auto.py verwenden.", file=sys.stderr)
    config = load_config()
    excel_path = _HERE / config["excel_file"]
    mappings: list[dict] = config["mappings"]

    bestand_mappings = [(m["isbn"], m["bestand_cell"]) for m in mappings if "bestand_cell" in m]
    angemeldet_mappings = [(m["isbn"], m["angemeldet_cell"]) for m in mappings if "angemeldet_cell" in m]
    unique_isbns = {isbn for isbn, _ in bestand_mappings}

    print(f"Verbinde mit IServ ({os.environ.get('ISERV_DOMAIN', '?')})...")
    client = AusleiheClient()
    print(f"Lade {len(unique_isbns)} Serien von der API...")

    series_total: dict[str, int] = {}
    series_title: dict[str, str] = {}
    retrieval_errors: list[str] = []
    for isbn in sorted(unique_isbns):
        try:
            s = client.series.get_by_isbn(isbn)
            if s.total is not None:
                series_total[isbn] = s.total
                series_title[isbn] = s.title
            else:
                print(f"  WARNUNG: {isbn} — 'total' fehlt in API-Antwort")
        except NotFoundError:
            print(f"  WARNUNG: {isbn} — nicht gefunden (abgelöste Serie?)")
        except Exception as e:
            print(f"  FEHLER:   {isbn} — {e}")
            retrieval_errors.append(f"{isbn}: {e}")

    if retrieval_errors and not args.allow_partial:
        raise SystemExit("API-Abruf unvollständig; keine Excel-Datei wird verändert. --allow-partial nur bewusst verwenden.")

    wb = load_workbook(str(excel_path))
    ws = wb[config["sheet_name"]]

    changed: list[tuple[str, object, int, str]] = []
    unchanged: int = 0
    not_found: list[str] = []
    a_changed: list[tuple[str, object, int]] = []  # vorinitialisiert für den finalen Save-Check

    for isbn, cell in bestand_mappings:
        if isbn not in series_total:
            not_found.append(f"{cell} (ISBN {isbn})")
            continue
        anchor = resolve_anchor(ws, cell)
        old = ws[anchor].value
        new = series_total[isbn]
        if old != new:
            ws[anchor] = new
            changed.append((cell, old, new, series_title[isbn]))
        else:
            unchanged += 1

    print()
    if args.dry_run:
        print("-- DRY RUN: keine Datei wird gespeichert --")

    if changed:
        print(f"{len(changed)} Zelle(n) {'würden aktualisiert' if args.dry_run else 'aktualisiert'}:")
        for cell, old, new, title in changed:
            print(f"  {cell}: {old} -> {new}  [{title}]")
    else:
        print("Keine Änderungen.")

    if unchanged:
        print(f"{unchanged} Zelle(n) bereits aktuell.")

    if not_found:
        print(f"\n{len(not_found)} Zelle(n) übersprungen (ISBN nicht in API):")
        for meldung in not_found:
            print(f"  {meldung}")

    if angemeldet_mappings:
        schoolyear_id = pick_schoolyear(client, args.schoolyear)
        print(f"\nLade Anmeldungen für Schuljahr {schoolyear_id}...")
        try:
            enrollment_counts = fetch_enrollment_counts(client, schoolyear_id)
        except Exception as e:
            print(f"FEHLER beim Laden der Anmeldungen: {e}")
            if not args.allow_partial:
                raise SystemExit(
                    "Anmeldungen konnten nicht vollständig geladen werden; keine Excel-Datei wird verändert. "
                    "--allow-partial nur bewusst verwenden."
                ) from e
            enrollment_counts = None

        if enrollment_counts is not None:
            # Ohne Annotation: die steht schon an der Vorinitialisierung oben,
            # und ein zweites Mal waere es eine Neudeklaration derselben
            # Variablen im selben Funktionsraum.
            a_changed = []
            a_unchanged = 0
            a_zero: list[str] = []
            for isbn, cell in angemeldet_mappings:
                anchor = resolve_anchor(ws, cell)
                new = enrollment_counts.get(isbn, 0)
                old = ws[anchor].value
                try:
                    same = old is not None and int(old) == new
                except (TypeError, ValueError):
                    same = False
                if not same:
                    ws[anchor] = new
                    a_changed.append((cell, old, new))
                else:
                    a_unchanged += 1
                if new == 0:
                    a_zero.append(f"{cell} (ISBN {isbn})")

            if a_changed:
                print(f"{len(a_changed)} 'Angemeldet'-Zelle(n) {'würden aktualisiert' if args.dry_run else 'aktualisiert'}:")
                for cell, old, new in a_changed:
                    print(f"  {cell}: {old} -> {new}")
            else:
                print("Keine Änderungen bei 'Angemeldet'-Zellen.")

            if a_unchanged:
                print(f"{a_unchanged} 'Angemeldet'-Zelle(n) bereits aktuell.")

            if a_zero:
                print(f"\n{len(a_zero)} 'Angemeldet'-Zelle(n) mit 0 Anmeldungen (keine Treffer im Schuljahr):")
                for meldung in a_zero:
                    print(f"  {meldung}")

    # Einmal speichern, nachdem alle Zellen (Bestand + Angemeldet) gesetzt wurden.
    if changed or a_changed:
        if not args.dry_run:
            backup = None if args.no_backup else excel_path.parent / "backups"
            backup_path = atomic_save_workbook(wb, excel_path, backup_dir=backup)
            print(f"Gespeichert: {excel_path}")
            if backup_path:
                print(f"Backup: {backup_path}")


if __name__ == "__main__":
    main()
