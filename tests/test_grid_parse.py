"""Struktur-Parser: Anker, Mehrjahresbänder, Fach-Fallback, Sperrflächen."""
from __future__ import annotations

from openpyxl import load_workbook

from bestand.core import parse_grid, resolve_anchor, strip_hint
from bestand.core.grid import SKIP_NOT_OFFERED, SKIP_ZUSTAND_NOT_WRITABLE, find_blocks


def _grid(workbook_path, **kwargs):
    wb = load_workbook(str(workbook_path))
    ws = wb["Bestand- und Nachbestellung"]
    return ws, parse_grid(ws, **kwargs)


def test_blocks_are_four_columns_wide(workbook_path):
    _, grid = _grid(workbook_path)
    assert grid.blocks == ((2, 5), (6, 9), (10, 13))


def test_stand_row_found(workbook_path):
    _, grid = _grid(workbook_path)
    assert grid.stand_rows == (10,)


def test_multi_year_band_groups_by_bestand_merge(workbook_path):
    """Erdkunde G3:G4 fasst Jg 5 und 6 zusammen; Angemeldet bleibt je Jahrgang."""
    _, grid = _grid(workbook_path)
    band = next(e for e in grid.entries if e.slots["bestand"].ref == "G3")
    assert band.grades == (5, 6)
    assert band.grade_label == "5-6"
    assert band.angemeldet_refs == ("F3", "F4")
    assert band.slots["zu_bestellen"].ref == "I3"


def test_single_year_entries_are_separate(workbook_path):
    """Deutsch hat keine Verbünde - jeder Jahrgang eine eigene Zeile."""
    _, grid = _grid(workbook_path)
    deutsch = [e for e in grid.entries if e.fach_label == "Deutsch"]
    assert [e.grades for e in deutsch] == [(5,), (6,), (7,), (12,)]
    assert [e.slots["bestand"].ref for e in deutsch] == ["C3", "C4", "C5", "C8"]


def test_fach_falls_back_to_higher_row(workbook_path):
    """B7:E7 ist leer -> die Oberstufenzeile erbt 'Deutsch' aus Zeile 1."""
    _, grid = _grid(workbook_path)
    entry = next(e for e in grid.entries if e.slots["bestand"].ref == "C8")
    assert entry.fach_label == "Deutsch"


def test_lower_fach_row_does_not_leak_upwards(workbook_path):
    """Zeile 7 ('Erdkunde (eA)') darf die Jahrgänge 5-7 darüber nicht umdeuten."""
    _, grid = _grid(workbook_path)
    labels = {c.fach_label for c in grid.cells if c.col == 6 and c.row in (3, 4, 5)}
    assert labels == {"Erdkunde"}
    assert next(c for c in grid.cells if c.col == 6 and c.row == 8).fach_label == "Erdkunde (eA)"


def test_hint_split(workbook_path):
    _, grid = _grid(workbook_path)
    cell = next(c for c in grid.cells if c.col == 6 and c.row == 8)
    assert (cell.subject, cell.hint) == ("Erdkunde", "eA")
    assert strip_hint("Deutsch") == ("Deutsch", None)


def test_zu_bestellen_column_is_not_writable(workbook_path):
    _, grid = _grid(workbook_path)
    cell = next(c for c in grid.cells if c.col == 5 and c.row == 3)   # Spalte E
    assert cell.skip_reason == SKIP_ZUSTAND_NOT_WRITABLE


def test_blocked_areas_are_full_block_width_merges(workbook_path):
    """J3:M4 und J8:M8 sperren Latein; nur Jg 7 bleibt eine echte Lücke."""
    _, grid = _grid(workbook_path)
    assert grid.blocked == ("J3:M4", "J8:M8")
    blocked_cells = {(c.row, c.col) for c in grid.cells if c.skip_reason == SKIP_NOT_OFFERED}
    assert (3, 10) in blocked_cells and (4, 12) in blocked_cells and (8, 10) in blocked_cells
    assert (5, 10) not in blocked_cells
    # Jg 5/6 und 12 sind gesperrt; die Zeile fuer Jg 7 bleibt bestehen, dort
    # fehlt nur das Buch in der Buecherliste.
    latein = [e for e in grid.entries if e.fach_label == "Latein"]
    assert [e.grades for e in latein] == [(7,)]


def test_skip_blocked_off_restores_old_behaviour(workbook_path):
    _, grid = _grid(workbook_path, skip_blocked=False)
    assert grid.blocked == ("J3:M4", "J8:M8")   # erkannt, aber nicht angewandt
    assert not any(c.skip_reason == SKIP_NOT_OFFERED for c in grid.cells)


def test_resolve_anchor_returns_top_left(workbook_path):
    ws, _ = _grid(workbook_path)
    assert resolve_anchor(ws, 4, 7) == (3, 7)     # G4 -> G3
    assert resolve_anchor(ws, 5, 7) == (5, 7)     # unverbunden


def test_find_blocks_ignores_unlabelled_columns(workbook_path):
    ws, grid = _grid(workbook_path)
    assert find_blocks(ws, list(grid.zustand_rows), ws.max_column) == list(grid.blocks)
