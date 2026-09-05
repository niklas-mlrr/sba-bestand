"""apply_snapshot: richtige Anker, Dedup-Regeln, Diagnosen statt Abbruch."""
from __future__ import annotations

from datetime import datetime

import pytest
from conftest import ISBN_DEUTSCH_5, ISBN_ERDKUNDE_56, SHEET_NAME
from openpyxl import load_workbook

from bestand.core import (
    BestandConfig,
    Snapshot,
    UpdateResult,
    apply_snapshot,
    fetch_snapshot,
    load_bestellt_counts,
    parse_grid,
    write_stand,
)


@pytest.fixture()
def prepared(workbook_path, fake_client):
    wb = load_workbook(str(workbook_path))
    ws = wb[SHEET_NAME]
    grid = parse_grid(ws)
    snapshot = fetch_snapshot(fake_client, "2026/2027",
                             fetched_at=datetime(2026, 9, 4, 12, 0, 0))
    counts, errors = load_bestellt_counts(wb["bestellt"])
    config = BestandConfig(excel_path=workbook_path, sheet_name=SHEET_NAME, safety_stock=5)
    return wb, ws, grid, snapshot, counts, errors, config


def _apply(prepared, **kwargs):
    wb, ws, grid, snapshot, counts, errors, config = prepared
    result = apply_snapshot(ws, grid, snapshot, config, bestellt_counts=counts,
                            result=UpdateResult(diagnostics=list(errors)), **kwargs)
    return wb, ws, grid, snapshot, result


def test_writes_land_on_anchor_cells(prepared):
    _, ws, _, _, result = _apply(prepared)
    assert ws["B3"].value == 50    # Deutsch Jg 5, Angemeldet
    assert ws["C3"].value == 40    # Bestand aus den Serien-Daten
    assert ws["G3"].value == 60    # Erdkunde-Band, Anker der Bestand-Zelle
    assert ws["G4"].value is None  # Verbund: nur der Anker wird geschrieben
    assert result.ok


def test_angemeldet_stays_per_grade_in_a_band(prepared):
    _, ws, _, _, _ = _apply(prepared)
    assert (ws["F3"].value, ws["F4"].value) == (48, 44)


def test_bestellt_bleibt_stehen_wenn_die_isbn_im_blatt_fehlt(prepared):
    """Fehlt die ISBN im Blatt 'bestellt', wird die Zelle NICHT angefasst.

    "Bestellt" ist die einzige Spalte, die auch von Hand gepflegt wird - und
    bis 2026-09-05 leerte der Abruf sie genau dann, wenn im Blatt 'bestellt'
    nichts zu der ISBN stand. Ein fehlender Eintrag dort heisst aber
    "unbekannt", nicht "nichts bestellt".
    """
    _, ws, _, _, _, _, _ = prepared
    ws["D5"] = 7    # von Hand eingetragen, ISBN steht nicht im Blatt 'bestellt'
    _apply(prepared)
    assert ws["D4"].value == 15    # 10 + 5 aus zwei Bestellzeilen
    assert ws["D5"].value == 7


def test_ausgelassene_bestellt_zelle_ist_keine_aenderung(prepared):
    """Was nicht geschrieben wird, darf auch nicht als 'aktualisiert' zaehlen."""
    _, _, _, _, result = _apply(prepared)
    bestellt_refs = {c.ref for c in result.changes if c.note.endswith("Bestellt")}
    assert "D4" in bestellt_refs      # steht im Blatt 'bestellt' -> geschrieben
    assert "D5" not in bestellt_refs  # fehlt dort -> ausgelassen


def test_ausgelassene_bestellt_zelle_geht_trotzdem_in_den_bedarf_ein(prepared):
    """Der stehen gelassene Wert wird gerechnet, nicht ignoriert.

    Sonst waere Angemeldet - Bestand - Bestellt zu gross und das Blatt
    'zu Bestellen' bestellte dieselben Buecher ein zweites Mal.
    """
    _, ws, _, _, _, _, _ = prepared
    ws["D5"] = 7
    _, _, _, _, result = _apply(prepared)
    eintrag = next(e for e in result.zu_bestellen_data.values()
                   if e["title"] == "Deutschbuch 7")
    assert eintrag["bestellt"] == 7


def test_text_in_der_bestellt_zelle_haelt_die_bedarfsrechnung_nicht_an(prepared):
    """Ein Formelrest oder eine Notiz zaehlt als leer, statt einen TypeError zu werfen."""
    _, ws, _, _, _, _, _ = prepared
    ws["D5"] = "noch offen"
    _, _, _, _, result = _apply(prepared)
    eintrag = next(e for e in result.zu_bestellen_data.values()
                   if e["title"] == "Deutschbuch 7")
    assert eintrag["bestellt"] is None
    assert ws["D5"].value == "noch offen"


def test_formula_cells_are_untouched(prepared):
    """Die Spalte 'zu bestellen' bleibt Formel - nie ein gerechneter Wert."""
    _, ws, _, _, _ = _apply(prepared)
    assert ws["E3"].value == "=B3-C3-D3"
    assert ws["I3"].value == "=F3+F4-G3-H3"


def test_blocked_area_is_never_written(prepared):
    _, ws, _, _, result = _apply(prepared)
    for ref in ("J3", "K3", "L3", "J4", "J8", "K8"):
        assert ws[ref].value is None
    assert not any(line.startswith("Sp.J/Zeile 3") for line in result.skipped)


def test_missing_book_is_skipped_not_fatal(prepared):
    _, _, _, _, result = _apply(prepared)
    assert result.ok
    assert [line for line in result.skipped] == [
        "Sp.J/Zeile 5: Kein Buch-Match für Fach 'Latein'.",
        "Sp.K/Zeile 5: Kein Buch-Match für Fach 'Latein'.",
        "Sp.L/Zeile 5: Kein Buch-Match für Fach 'Latein'.",
    ]


def test_ambiguous_match_becomes_diagnostic(workbook_path, fake_client):
    """Zwei Bücher für dasselbe Fach -> Diagnose, der Aufrufer darf nicht speichern."""
    wb = load_workbook(str(workbook_path))
    ws = wb[SHEET_NAME]
    grid = parse_grid(ws)
    snapshot = fetch_snapshot(fake_client, "2026/2027", eager=True)
    books = dict(snapshot.grade_books)
    books[5] = books[5] + [{"isbn": "9783128640204", "title": "Zweitbuch Deutsch",
                            "subjects": ["Deutsch"]}]
    snapshot = Snapshot(
        schoolyear_id=snapshot.schoolyear_id, fetched_at=snapshot.fetched_at,
        booklists_by_grade=snapshot.booklists_by_grade, grade_books=books,
        enrolled=snapshot.enrolled, paid=snapshot.paid, series_data=snapshot.series_data,
    )
    config = BestandConfig(excel_path=workbook_path, sheet_name=SHEET_NAME)
    result = apply_snapshot(ws, grid, snapshot, config)
    assert not result.ok
    assert any("Mehrdeutiger Buch-Match" in line for line in result.diagnostics)


def test_override_resolves_ambiguity(workbook_path, fake_client):
    wb = load_workbook(str(workbook_path))
    ws = wb[SHEET_NAME]
    grid = parse_grid(ws)
    snapshot = fetch_snapshot(fake_client, "2026/2027", eager=True)
    books = dict(snapshot.grade_books)
    books[5] = books[5] + [{"isbn": "9783128640204", "title": "Zweitbuch Deutsch",
                            "subjects": ["Deutsch"]}]
    snapshot = Snapshot(
        schoolyear_id=snapshot.schoolyear_id, fetched_at=snapshot.fetched_at,
        booklists_by_grade=snapshot.booklists_by_grade, grade_books=books,
        enrolled=snapshot.enrolled, paid=snapshot.paid, series_data=snapshot.series_data,
    )
    config = BestandConfig(excel_path=workbook_path, sheet_name=SHEET_NAME,
                           match_overrides={"5|Deutsch|": ISBN_DEUTSCH_5})
    result = apply_snapshot(ws, grid, snapshot, config)
    assert result.ok
    assert ws["B3"].value == 50


def test_bestellt_sheet_errors_come_first(prepared):
    wb, ws, grid, snapshot, counts, _errors, config = prepared
    wb["bestellt"]["C2"] = "keine Zahl"
    counts, errors = load_bestellt_counts(wb["bestellt"])
    result = apply_snapshot(ws, grid, snapshot, config, bestellt_counts=counts,
                            result=UpdateResult(diagnostics=list(errors)))
    assert result.diagnostics[0].startswith("bestellt!2:")


def test_isbn_by_entry_maps_grid_keys(prepared):
    _, _, grid, _, result = _apply(prepared)
    band = next(e for e in grid.entries if e.slots["bestand"].ref == "G3")
    assert result.isbn_by_entry[band.key] == ISBN_ERDKUNDE_56


def test_write_stand_sets_datetime_and_format(prepared):
    _, ws, grid, _, result = _apply(prepared)
    when = datetime(2026, 9, 4, 12, 0, 0)
    write_stand(ws, grid, when, result)
    assert ws["B10"].value == when
    assert ws["B10"].number_format.startswith("dddd")
    assert result.changes[-1].render().endswith("[Stand/Abfragezeitpunkt]")


def test_lazy_grade_books_load_once(workbook_path, fake_client):
    """Jede Bücherliste wird höchstens einmal geholt."""
    events: list[tuple[str, dict]] = []
    snapshot = fetch_snapshot(fake_client, "2026/2027",
                              progress=lambda ev, p: events.append((ev, p)))
    wb = load_workbook(str(workbook_path))
    ws = wb[SHEET_NAME]
    grid = parse_grid(ws)
    config = BestandConfig(excel_path=workbook_path, sheet_name=SHEET_NAME)
    apply_snapshot(ws, grid, snapshot, config)
    grades = [p["grade"] for ev, p in events if ev == "grade_books"]
    assert grades == [5, 6, 7, 12]


def test_enrollment_counts_ignore_deleted(fake_client):
    snapshot = fetch_snapshot(fake_client, "2026/2027")
    assert snapshot.enrolled[(5, ISBN_DEUTSCH_5)] == 50
    assert snapshot.paid[(5, ISBN_DEUTSCH_5)] == 40


def test_ebooks_are_filtered_out(fake_client):
    snapshot = fetch_snapshot(fake_client, "2026/2027", eager=True)
    titles = [book["title"] for book in snapshot.grade_books[5]]
    assert "Deutschbuch eBook" not in titles
