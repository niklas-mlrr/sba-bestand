"""Blatt 'zu Bestellen': Geometrie, Bestell-Nr.-Übernahme, Ergebniszeile."""
from __future__ import annotations

from datetime import datetime

import pytest
from conftest import SHEET_NAME
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from bestand.core import (
    BestandConfig,
    UpdateResult,
    apply_snapshot,
    compute_zu_bestellen_rows,
    fetch_snapshot,
    load_bestellt_counts,
    parse_grid,
    rebuild_zu_bestellen,
)


@pytest.fixture()
def rebuilt(workbook_path, fake_client):
    wb = load_workbook(str(workbook_path))
    ws = wb[SHEET_NAME]
    grid = parse_grid(ws)
    snapshot = fetch_snapshot(fake_client, "2026/2027",
                              fetched_at=datetime(2026, 9, 4, 12, 0, 0))
    counts, errors = load_bestellt_counts(wb["bestellt"])
    config = BestandConfig(excel_path=workbook_path, sheet_name=SHEET_NAME, safety_stock=5)
    result = apply_snapshot(ws, grid, snapshot, config, bestellt_counts=counts,
                            result=UpdateResult(diagnostics=list(errors)))
    rows = rebuild_zu_bestellen(wb, result, snapshot, config.safety_stock)
    return wb, result, rows


def test_demand_is_computed_in_python(rebuilt):
    """Bedarf = Angemeldet - Bestand - Bestellt, nie aus der Excel-Formel."""
    _, _, rows = rebuilt
    erdkunde = next(r for r in rows if r.title == "Terra 5/6")
    # 48 + 44 angemeldet, 60 im Bestand, 30 bestellt -> 2, plus 5 Sicherheit
    assert erdkunde.stueckzahl == 7
    assert erdkunde.grade == 5


def test_rows_sorted_by_title(rebuilt):
    _, _, rows = rebuilt
    assert [r.title for r in rows] == sorted(r.title for r in rows)


def test_only_positive_demand_is_listed(rebuilt):
    """Deutschbuch 7 hat 90 Exemplare bei 52 Anmeldungen - kein Bedarf."""
    _, _, rows = rebuilt
    assert "Deutschbuch 7" not in [r.title for r in rows]


def test_bestellnr_follows_the_isbn(rebuilt):
    """Die alte Bestell-Nr. 42 stand bei Terra 5/6 und muss dort wieder landen."""
    wb, _, rows = rebuilt
    ws = wb["zu Bestellen"]
    terra_row = next(r for r in range(2, ws.max_row + 1)
                     if ws.cell(r, 5).value == "Terra 5/6")
    assert ws.cell(terra_row, 1).value == 42


def test_table_geometry_matches_row_count(rebuilt):
    wb, _, rows = rebuilt
    ws = wb["zu Bestellen"]
    table = ws.tables["zuBestellen"]
    _, header_row, _, last_row = range_boundaries(table.ref)
    assert header_row == 1
    assert last_row == 1 + len(rows) + 1        # Kopf + Daten + Ergebniszeile
    _, _, _, filter_last = range_boundaries(table.autoFilter.ref)
    assert filter_last == 1 + len(rows)
    assert table.sortState is None


def test_totals_row_rebuilt(rebuilt):
    wb, _, rows = rebuilt
    ws = wb["zu Bestellen"]
    totals_row = 1 + len(rows) + 1
    assert ws.cell(totals_row, 1).value == "Ergebnis"
    assert ws.cell(totals_row, 4).value == "=SUBTOTAL(109,zuBestellen[Stückzahl])"
    assert ws.cell(totals_row, 9).value == "=SUBTOTAL(109,zuBestellen[Gesamtpreis (brutto)])"


def test_gesamtpreis_formula_in_every_data_row(rebuilt):
    wb, _, rows = rebuilt
    ws = wb["zu Bestellen"]
    for row in range(2, 2 + len(rows)):
        assert ws.cell(row, 9).value.startswith("=IF(OR(zuBestellen[[#This Row],[Stückzahl]]")


def test_no_demand_keeps_one_empty_data_row(workbook_path, fake_client):
    """Excel verlangt mindestens eine Datenzeile; die Ergebniszeile rutscht nach."""
    wb = load_workbook(str(workbook_path))
    snapshot = fetch_snapshot(fake_client, "2026/2027")
    result = UpdateResult()
    rows = rebuild_zu_bestellen(wb, result, snapshot, 5)
    ws = wb["zu Bestellen"]
    assert rows == []
    _, _, _, last_row = range_boundaries(ws.tables["zuBestellen"].ref)
    assert last_row == 3
    assert ws.cell(2, 5).value is None
    assert ws.cell(3, 1).value == "Ergebnis"


def test_compute_rows_without_workbook():
    """Die Rechnung selbst braucht kein Excel."""
    data = {"9780000000015": {"angemeldet": 10, "bestand": 3, "bestellt": 2,
                              "title": "Titel", "grades": {5, 6}, "fach": "Deutsch"}}
    series = {"9780000000015": {"title": "Serientitel", "publisher": "Verlag",
                                "price": 9.5, "total": 3}}
    rows = compute_zu_bestellen_rows(data, series, safety_stock=4)
    assert len(rows) == 1
    assert (rows[0].stueckzahl, rows[0].grade, rows[0].title) == (9, 5, "Serientitel")
